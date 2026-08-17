"""Excel-Export mit Kopfzeile, Autofilter und klickbaren URLs."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from fbgroups.export.columns import COLUMNS, headers, row_values
from fbgroups.models import Group

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_LINK_FONT = Font(color="0563C1", underline="single")

_COLUMN_WIDTHS = {
    "Score": 8, "Score Reason": 46, "Status": 18, "Validation Status": 17,
    "Data Quality": 13, "Gruppenname": 42, "URL": 46, "Zielgruppen": 18,
    "Stadt": 14, "Bundesland": 22, "Kategorie": 18, "Mitglieder (ca.)": 16,
    "Sichtbarkeit": 13, "Notizen": 30,
    "Konf. Zielgruppe": 16, "Konf. Stadt": 12, "Funde": 8, "Group-ID": 24,
}

# Farbliche Hervorhebung der Statuswerte - die Sichtung soll auf einen Blick
# erkennen, was brauchbar ist und was nachgepflegt werden muss.
_STATUS_FILLS = {
    "validated": PatternFill("solid", fgColor="E2EFDA"),
    "duplicate": PatternFill("solid", fgColor="FFF2CC"),
    "insufficient_data": PatternFill("solid", fgColor="FCE4D6"),
    "invalid": PatternFill("solid", fgColor="F8CBAD"),
}


def export_excel(groups: list[Group], path: Path, sheet_title: str = "Gruppen") -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_labels = headers()
    ws.append(header_labels)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    url_index = next(i for i, (field, _) in enumerate(COLUMNS, start=1) if field == "url_canonical")
    status_index = next(i for i, (field, _) in enumerate(COLUMNS, start=1) if field == "status")

    for group in groups:
        ws.append(row_values(group))
        cell = ws.cell(row=ws.max_row, column=url_index)
        cell.hyperlink = group.url_canonical
        cell.font = _LINK_FONT

        fill = _STATUS_FILLS.get(group.status.value)
        if fill is not None:
            ws.cell(row=ws.max_row, column=status_index).fill = fill

    for index, label in enumerate(header_labels, start=1):
        ws.column_dimensions[get_column_letter(index)].width = _COLUMN_WIDTHS.get(label, 16)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header_labels))}{ws.max_row}"

    wb.save(path)
    return path
