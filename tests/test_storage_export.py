from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from fbgroups.export import export_csv, export_excel
from fbgroups.models import (
    DataQuality,
    Group,
    ImportRun,
    Provenance,
    RecordStatus,
    SourceType,
)
from fbgroups.storage import SqliteStore

REAL_ID_A = "482910573829104"
REAL_ID_B = "739201847362915"


def make_group(group_id: str = REAL_ID_A, name: str = "Syrer in Berlin") -> Group:
    return Group(
        group_id=group_id,
        url_canonical=f"https://www.facebook.com/groups/{group_id}",
        name=name,
        audience_tags=["syrians"],
        audience_confidence=1.0,
        city="Berlin",
        bundesland="Berlin",
        city_confidence=1.0,
        category="community",
        member_count_hint=12000,
        score=77.5,
        score_reason="berechnet aus: audience_match, category_match, city_match, member_count",
        data_quality=DataQuality.COMPLETE,
        status=RecordStatus.VALIDATED,
        sources=[Provenance(source_type=SourceType.MANUAL_SEED, source_ref="test.csv")],
    )


def make_unscored_group(group_id: str = REAL_ID_B) -> Group:
    """Gruppe ohne Metadaten - Score bleibt None."""
    return Group(
        group_id=group_id,
        url_canonical=f"https://www.facebook.com/groups/{group_id}",
        score=None,
        score_reason="insufficient_data: kein Gruppenname vorhanden",
        data_quality=DataQuality.NONE,
        status=RecordStatus.INSUFFICIENT_DATA,
    )


def test_speichern_und_laden(tmp_path: Path) -> None:
    with SqliteStore(tmp_path / "test.sqlite") as store:
        new, known = store.upsert_groups([make_group()])
        assert (new, known) == (1, 0)

        loaded = store.load_groups()
        assert len(loaded) == 1
        assert loaded[0].name == "Syrer in Berlin"
        assert loaded[0].audience_tags == ["syrians"]
        assert loaded[0].city == "Berlin"


def test_erneuter_import_zaehlt_als_bekannt(tmp_path: Path) -> None:
    with SqliteStore(tmp_path / "test.sqlite") as store:
        store.upsert_groups([make_group()])
        new, known = store.upsert_groups([make_group()])

        assert (new, known) == (0, 1)
        assert store.count_groups() == 1
        assert store.load_groups()[0].times_seen == 2


def test_manuelle_notiz_bleibt_erhalten(tmp_path: Path) -> None:
    """Ein erneuter Import darf eine manuelle Notiz nicht zuruecksetzen."""
    db = tmp_path / "test.sqlite"
    with SqliteStore(db) as store:
        store.upsert_groups([make_group()])
        store.conn.execute("UPDATE groups SET notes='geprueft'")
        store.conn.commit()

    with SqliteStore(db) as store:
        store.upsert_groups([make_group()])
        assert store.load_groups()[0].notes == "geprueft"


def test_score_none_ueberlebt_die_datenbank(tmp_path: Path) -> None:
    """NULL darf beim Speichern nicht zu 0.0 werden."""
    with SqliteStore(tmp_path / "test.sqlite") as store:
        store.upsert_groups([make_unscored_group()])
        reloaded = store.load_groups()[0]

        assert reloaded.score is None
        assert reloaded.status is RecordStatus.INSUFFICIENT_DATA
        assert reloaded.data_quality is DataQuality.NONE
        assert "insufficient_data" in reloaded.score_reason


def test_bewertete_gruppen_stehen_vorne(tmp_path: Path) -> None:
    with SqliteStore(tmp_path / "test.sqlite") as store:
        store.upsert_groups([make_unscored_group(), make_group()])
        geladen = store.load_groups()

        assert geladen[0].score == 77.5
        assert geladen[-1].score is None


def test_arabischer_name_ueberlebt_die_datenbank(tmp_path: Path) -> None:
    with SqliteStore(tmp_path / "test.sqlite") as store:
        store.upsert_groups([make_group(REAL_ID_B, "سوريين في برلين")])
        assert store.load_groups()[0].name == "سوريين في برلين"


def test_lauf_wird_gespeichert(tmp_path: Path) -> None:
    with SqliteStore(tmp_path / "test.sqlite") as store:
        store.save_run(ImportRun(run_id="run1", rows_total=10, rows_valid=8))
        row = store.conn.execute("SELECT * FROM runs WHERE run_id='run1'").fetchone()
        assert row["rows_valid"] == 8


def test_csv_export(tmp_path: Path) -> None:
    target = export_csv([make_group(REAL_ID_A, "سوريين في برلين")], tmp_path / "out.csv")
    content = target.read_text(encoding="utf-8-sig")

    assert "Gruppenname" in content
    assert "سوريين في برلين" in content
    assert f"https://www.facebook.com/groups/{REAL_ID_A}" in content


def test_export_enthaelt_neue_spalten(tmp_path: Path) -> None:
    target = export_excel([make_group()], tmp_path / "out.xlsx")
    kopfzeile = [c.value for c in load_workbook(target).active[1]]

    for spalte in ("Validation Status", "Data Quality", "Score", "Score Reason", "Status"):
        assert spalte in kopfzeile, spalte


def test_export_zeigt_unknown_statt_leerer_zelle(tmp_path: Path) -> None:
    """Leere Zelle laesst offen, ob geprueft wurde - 'unknown' ist eindeutig."""
    target = export_excel([make_unscored_group()], tmp_path / "out.xlsx")
    ws = load_workbook(target).active
    kopfzeile = [c.value for c in ws[1]]

    def wert(spalte: str):
        return ws.cell(row=2, column=kopfzeile.index(spalte) + 1).value

    assert wert("Gruppenname") == "unknown"
    assert wert("Stadt") == "unknown"
    assert wert("Kategorie") == "unknown"
    assert wert("Mitglieder (ca.)") == "unknown"
    assert wert("Zielgruppen") == "unknown"
    # Score bleibt leer - kein Ersatzwert, keine Null
    assert wert("Score") in (None, "")
    assert "insufficient_data" in wert("Score Reason")


def test_excel_export(tmp_path: Path) -> None:
    groups = [make_group(), make_group(REAL_ID_B, "عرب في هامبورغ")]
    target = export_excel(groups, tmp_path / "out.xlsx")
    ws = load_workbook(target).active

    assert ws.max_row == 3           # Kopfzeile + zwei Gruppen
    assert ws.cell(row=1, column=1).value == "Score"
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None
    # URL-Spalte ist verlinkt
    url_col = [c.value for c in ws[1]].index("URL") + 1
    assert ws.cell(row=2, column=url_col).hyperlink is not None
